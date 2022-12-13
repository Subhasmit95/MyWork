# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import math
import random
#import blackboard.py


eta=0.5
momentum=0.8
lambda_val=0.8
Finalweights=[]
class Neuron:
    #initializing the variables in the __init__ function
    def __init__(self,activation_value,previous_layer,current_neuron,weight,deltaweight=[],localgradient=0): #previous_layer is not required in this class
            self.activation_value=activation_value
            self.previous_layer=previous_layer
            self.current_neuron=current_neuron
            self.weight=weight
            self.deltaweight=deltaweight
            self.localgradient=localgradient



    #The activation function
    def sigmoidcalc(self,r):
        sigmoid_val = 1.0/ (1.0 + math.exp(-(lambda_val * r)))
        return sigmoid_val

    #precalculation for generating activationvalue * weight
    def weightmultiplication(self ,previous_layer,current_neuron):
        #The multipliction
        if len(previous_layer)>0:
            res=0.0
            for i1 in range(len(previous_layer)):
                print("Length Previous: "+str(len(previous_layer)))
                print("i: " + str(i1))
                print('previous layer activation value',previous_layer[i1].activation_value,'previos layer weight neurom{} -'.format(current_neuron),self.previous_layer[i1].weight[current_neuron])
                res = res + (previous_layer[i1].activation_value * previous_layer[i1].weight[current_neuron])
                print(res)

            sigmoid = self.sigmoidcalc(res)
            print('The Sigmoid value is ',sigmoid)
            self.activation_value = sigmoid
            print('Activation value',self.activation_value)
        else:
            self.activation_value=activation_value
            print(self.activation_value)




def FeedForward(input_x,input_y):
    print('IN FEEDFORWARD : ')

    inputlayer[0].activation_value=float(input_x)
    inputlayer[1].activation_value=float(input_y)
    print('Total length using the bias - ',len(inputlayer))

    for x in range(len(inputlayer)):
        print('the neuron{}'.format(x))


    for h in range(int(hiddenlayersize - 1)):
        # hidden_layer.append(Neuron(0.0,inputlayer,h,[]))
        hidden_layer[h].weightmultiplication(inputlayer, h)
        print(hidden_layer[h].activation_value)

    for h in range(len(hidden_layer)):
        print('the activation of the last hidden layer ',hidden_layer[h].activation_value)
                                          
    # creating output layer
    for x in range(len(hidden_layer)):
        print('the neuron{}'.format(x))

    for out in range(int(len(output_layer))):
        #output_layer.append(Neuron(0.0, hidden_layer, o, []))
        output_layer[out].weightmultiplication(hidden_layer, out)
        print('The final output - ',output_layer[out].activation_value)

#The error calculation
def errorcalculation(output_value):
    valerror = []
    valsumoferrors = 0.0
    valsqaureoferrors = 0.0
    # Initialising the errors
    for e in range(len(output_value)):
        valerror.append(0.0)
        print("Output Values: "+str(output_value[e]))

    # finding the error in case of the the output noodes

    for out in range(len(output_layer)):
        valerror[out] = output_value[out] - output_layer[out].activation_value
        # finding the error in case of the the output noodes

    valsumoferrors = sum(valerror)
    print('sum of errors validation - ', valsumoferrors)
    valsqaureoferrors = (valsumoferrors ** 2) / len(valerror)
    print('Square of errors in validation - ', valsqaureoferrors)
    return valsqaureoferrors


#BACKPROPAGATION
def backpropagation(output_val):
    print('IN BACKPROPAGATION : ')
     #the error list for the actual value - calculated value
    error=[]
    sumoferrors=0.0
    sqaureoferrors=0.0
    #Initialising the errors
    for e in range(len(output_val)):
        error.append(0.0)

     #finding the error in case of the the output noodes

    for out in range(len(output_layer)):
        error[out]=output_val[out] - output_layer[out].activation_value

    sumoferrors = sum(error)
    print('sum of errors - ',sumoferrors)
    sqaureoferrors = (sumoferrors**2)/len(error)
    print('Square of errors - ', sqaureoferrors)

    # finding the local gradient of each output layer


    for out in range(len(output_layer)):
        output_layer[out].localgradient = lambda_val * output_layer[out].activation_value * (1.0-output_layer[out].activation_value)*error[out]




    print('the length of hidden layer ',len(hidden_layer))
    # finding the local gradient of each hidden layer
    for h in range(len(hidden_layer)):
        sumofweightsandgradient=0.0
        for out in range(len(output_layer)):
            # calculating the summation of the weight connected to each output node multiplied with the local gradient of that output node
            sumofweightsandgradient = sumofweightsandgradient + (hidden_layer[h].weight[out] * output_layer[out].localgradient)
        hidden_layer[h].localgradient = lambda_val * hidden_layer[h].activation_value * (1 - hidden_layer[h].activation_value ) * sumofweightsandgradient



    #Calculating the delta weights hidden layer
    for h in range(len(hidden_layer)):
        print("Hidden Layer Delta Weight: "+str(hidden_layer[h].deltaweight))
        for out in range(len(output_layer)):
            hidden_layer[h].deltaweight[out]=(eta * output_layer[out].localgradient * hidden_layer[h].activation_value) + (momentum * hidden_layer[h].deltaweight[out])

    # Calculating the delta weights inputlayer
    for i in range(len(inputlayer)):
         for h in range(len(hidden_layer)-1):
             inputlayer[i].deltaweight[h] = (eta * hidden_layer[h].localgradient * inputlayer[i].activation_value) + (momentum * inputlayer[i].deltaweight[h])

    #Updating the weights of the hidden layer
    for h in range(len(hidden_layer)):
         for out in range(len(output_layer)):
            hidden_layer[h].weight[out] = hidden_layer[h].weight[out] + hidden_layer[h].deltaweight[out]
            print('The Hidden layer weights : ',hidden_layer[h].weight[out], hidden_layer[h].deltaweight[out])

    #Updating the weights of the input layer
    for i in range(len(inputlayer)):
         for h in range(len(hidden_layer)-1):
             inputlayer[i].weight[h] = inputlayer[i].weight[h] + inputlayer[i].deltaweight[h]
             print('The Input layer weights : ', inputlayer[i].weight[h], inputlayer[i].deltaweight[h])

    return sqaureoferrors
from openpyxl import load_workbook

def trainingFunc():

    numberofepoch=1
    meansquareerror=0
    valmeansquareerror=0
    flag=True
    while flag:
        # Start of the Training set epoch
        print("############################### FIRST EPOCH{} ################################ \n \n ".format(numberofepoch))
        print('TRAINING PART')
        wb = load_workbook("mainxor.xlsx")
        sheet = wb['Sheet1']
        for r in range(sheet.max_row):
            totalerror = []

            inputs = []
            outputs = []
            for c in range(1, 3):
                inputs.append(sheet.cell(row=r + 1, column=c).value)
            for oc in range(3, 4):
                outputs.append(sheet.cell(row=r + 1, column=oc).value)
            FeedForward(inputs[0], inputs[1])
            rowerrors = backpropagation(outputs)
            totalerror.append(rowerrors)

        print('', math.sqrt(sum(totalerror) / len(totalerror)))
        meansquareerror = math.sqrt(sum(totalerror) / len(totalerror))
        print('THE TOTAL MEAN SQUARE OF THE TRAINING  ', meansquareerror)

        # End of the training epoch

        # Validation part
        print('THE VALIDATION PART ')
        wb = load_workbook("mainxor.xlsx")
        sheet = wb['Sheet1']
        for r in range(sheet.max_row):
            valtotalerror = []

            valinputs = []
            valoutputs = []

            for c in range(1, 3):
                valinputs.append(sheet.cell(row=r + 1, column=c).value)
            for oc in range(3, 4):
                valoutputs.append(sheet.cell(row=r + 1, column=oc).value)
            print('FeedForward of the validation')
            FeedForward(valinputs[0], valinputs[1])
            valtotalerror.append(errorcalculation(valoutputs))
        print('Total validation errors of 1 epoch', math.sqrt(sum(valtotalerror) / len(valtotalerror)))
        valmeansquareerror = math.sqrt(sum(valtotalerror) / len(valtotalerror))

        print('THE TOTAL MEAN SQUARE OF THE TRAINING  ', meansquareerror)
        print('THE TOTAL MEAN SQUARE OF THE validation set  ', valmeansquareerror)

        numberofepoch=numberofepoch+1
        print('Number of Epoch  ',numberofepoch)
        if meansquareerror < 0.1:
            flag=False

    #FeedForward(0.0,0.0)
    #backpropagation(validation[0])
def predictval():
    x=float(input('first input'))
    y = float(input('second input'))
    inputlayer[-1].activation_value = 1.0
    hidden_layer[-1].activation_value= 1.0
    print("The Hidden layers ", hidden_layer[-1].activation_value,hidden_layer[-1].weight[:])
    print("The input layers",inputlayer[2].activation_value,inputlayer[2].weight[:])
    FeedForward(x,y)

if __name__ == '__main__':



    #creating the input layer
    hiddenlayersize=int(input("HOW MANY NEURON YOU WANT IN THE HIDDEN LAYER \n "))
    outputsize=int(input('How many output do you want \n'))
    inputlayersize=2
    #declaring the layers
    inputlayer=[]
    hidden_layer = []
    output_layer = []
    #initializing the layers
    hiddenlayersize=hiddenlayersize+1     #adding bias value to the hidden layer
    inputlayersize=inputlayersize+1       #adding the bias value to the output layer
    print(hiddenlayersize,inputlayersize)
    #delta weight assignment for hidden layer

    #now


    for i in range(inputlayersize):
        inputlayer.append(Neuron(0.0,[],i,[]))
    print('Value of Input layer while initializing : ',inputlayer[0].deltaweight)

    inputlayer[-1].activation_value=1.0 #for the bias neuron

    for h in range(hiddenlayersize):
        hidden_layer.append(Neuron(0.0, inputlayer, h, []))

    hidden_layer[-1].activation_value=1.0 #for the bias neuron

    for o in range(outputsize):
        output_layer.append(Neuron(0.0, hidden_layer, o, []))
    #Generating random weights for Input to hidden layers

    for i in range(len(inputlayer)):
        for j in range(int(hiddenlayersize)-1):
            # wh=float(input('enter the hidden layer weight please\n'))
            inputlayer[i].weight.append(float(random.uniform(0,1)))  # float(random.random())
        print("Input Weights: " + str(inputlayer[i].weight))

    # creating weight for hidden to output layer

    for i in range(len(hidden_layer)):
        for j in range(int(outputsize)):
            # hwo = float(input('enter the hidden to output layer weight please\n'))
            hidden_layer[i].weight.append(float(random.uniform(0,1)))

    #calculating hidden layers delta weight
    for i in range(len(hidden_layer)):
        hidden_layer[i].deltaweight = []
        for j in range(len(output_layer)):
            # wh=float(input('enter the hidden layer weight please\n'))
            hidden_layer[i].deltaweight.append(0.0)

    # delta weight assignment for input layer
    for i in range(len(inputlayer)):
        inputlayer[i].deltaweight = []
        for j in range(len(hidden_layer)):
            # wh=float(input('enter the hidden layer weight please\n'))
            inputlayer[i].deltaweight.append(0.0)

    trainingFunc()

    predictval()
    predictval()
    predictval()
    predictval()
    predictval()
    predictval()
    #calling the feedforward process

    '''
    FeedForward(0.0,0.0)
    backpropagation([1,1])


#Rejected part

    #Creating the weight for the input to hidden layer

    for i in range(len(inputlayer)):
        initial_weight=[]
        for j in range(int(hiddenlayersize)):
                inputlayer[i].weight.append(float(random.random()))
        print("Input Weights: " +str(inputlayer[i].weight))

    

    #Creating Hidden Layer

    for x in range(len(inputlayer)):
        print('the neuron{}'.format(x))

    for h in range(int(hiddenlayersize-1)):
        #hidden_layer.append(Neuron(0.0,inputlayer,h,[]))
        hidden_layer[h].weightmultiplication(inputlayer,h)
        print(hidden_layer[h].activation_value)

    #creating weight for output layer
    for i in range(len(hidden_layer)):
        for j in range(int(outputsize)):
            hidden_layer[i].weight.append(float(random.random()))

    # creating output layer
    for x in range(len(hidden_layer)):
        print('the neuron{}'.format(x))

    for o in range(int(outputsize)):
        #output_layer.append(Neuron(0.0, hidden_layer, o, []))
        output_layer[o].weightmultiplication(hidden_layer, o)
        print(output_layer[o].activation_value)
'''

